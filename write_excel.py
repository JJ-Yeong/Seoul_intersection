import os
import glob
import math
import warnings
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
import seaborn as sns
import platform
from matplotlib import font_manager, rc
from datetime import date, datetime
from numpy import random
from pandas.core.frame import DataFrame as DF
from IPython.display import display
from openpyxl import Workbook, load_workbook
from typing import Union, List, Dict, Tuple

from build_df import preprocess_grouping, build_final_df
from build_df_config import PH_SPOT_MOVE_DICT, TXT_FILE
from write_excel_config import *

if platform.system()=="Windows":
    font_name=font_manager.FontProperties(fname="c:/Windows/Fonts/malgun.ttf").get_name()
    rc('font', family=font_name)
mpl.rcParams['axes.unicode_minus']=False




def get_fifteen_df(txt_file: Union[str, List[str]]) -> DF:
    if os.path.isdir(txt_file):
        path_list = glob.glob(os.path.join(txt_file, "*"))
        df = preprocess_grouping(path_list)
    elif os.path.splitext(txt_file)[-1] == ".txt":
        df = preprocess_grouping(txt_file)
    else:
        raise Exception("'txt_file'은 디렉토리 혹은 .txt 파일이어야 합니다!")

    final_df = build_final_df(df)
    final_df["지점"] = final_df["지점"].apply(lambda x: x.zfill(2))
    final_df["방향"] = final_df["방향"].apply(lambda x: x.zfill(2))
    final_df.sort_values(by=["지점", "방향", "차종", "h", "m"], inplace=True)
    # object_col = final_df.select_dtypes(['object'])
    # final_df[object_col.columns] = object_col.apply(lambda x: x.astype('category'))
    final_df = final_df.reset_index()
    agg_d = {
        "지점":"first", 
        "연월일":"first", 
        "일":"first", 
        "h":"first", 
        "m":"first", 
        "hm":"first", 
        "방향":"first", 
        "차종":"first", 
        "교통량":"sum"
        }
    fifteen_df = final_df.groupby(final_df.index // 15).agg(agg_d)
    fifteen_final_df = fifteen_df.pivot_table(index=["지점", "hm", "차종"], columns=["방향"], values="교통량", fill_value=0).reset_index()
    fifteen_final_df.rename_axis(None, axis=1, inplace=True) 

    return fifteen_final_df


def load_workbook_ignore_warning(input_path: str) -> Workbook:
    warnings.simplefilter("ignore")
    src_wb = load_workbook(filename=input_path)
    warnings.simplefilter("default")
    return src_wb


def get_random_value() -> float:
    # random_low = random.uniform(low=0.85, high=0.95, size=None)
    # random_high = random.uniform(low=1.05, high=1.15, size=None)
    # random_factor = random.uniform(low=random_low, high=random_high, size=None)
    random_factor = random.uniform(low=0.84, high=1.16, size=None)
    return random_factor


def read_and_compare_value(df: DF, input_path: str, output_path: str, sheet1: dict, sheet2: dict, movement_list: dict) -> Tuple[float, float]:
    src_wb: Workbook
    src_wb = load_workbook_ignore_warning(input_path=input_path)

    output_fig_path = output_path[:-5] + ".png"

    df.reset_index(drop=True, inplace=True)
    heatmap_df = df.copy()
    output_df = df.copy()

    i = 0

    src_ws1 = src_wb[sheet1["sheet_name"]]
    for row in sheet1["row_range1"]:
        for col in sheet1["col_range1"]:
            movement = col-2
            if movement in movement_list:
                movement_new_col = str(movement).zfill(2)
                movement_old_col = movement_new_col+"_2021"
                movement_diff_col = movement_new_col+"_diff"
                value_new = df.loc[i, movement_new_col]
                value_old = src_ws1.cell(row = row, column = col).value
                diff_ration = round((value_new - value_old)/value_old, 2) if value_old != 0 else 1.0
                heatmap_df.loc[i, movement_new_col] = diff_ration
                output_df.loc[i, movement_old_col] = value_old
                output_df.loc[i, movement_diff_col] = f"{int(diff_ration*100)}%"  
        i += 1

    for row in sheet1["row_range2"]:
        for col in sheet1["col_range2"]:
            movement = col-2
            if movement in movement_list:
                movement_new_col = str(movement).zfill(2)
                movement_old_col = movement_new_col+"_2021"
                movement_diff_col = movement_new_col+"_diff"
                value_new = df.loc[i, movement_new_col]
                value_old = src_ws1.cell(row = row, column = col).value
                diff_ration = round((value_new - value_old)/value_old, 2) if value_old != 0 else 1.0
                heatmap_df.loc[i, movement_new_col] = diff_ration        
                output_df.loc[i, movement_old_col] = value_old
                output_df.loc[i, movement_diff_col] = f"{int(diff_ration*100)}%"        
        i += 1

    src_ws2 = src_wb[sheet2["sheet_name"]]
    for row in sheet2["row_range1"]:
        for col in sheet2["col_range1"]:
            movement = col-2
            if movement in movement_list:
                movement_new_col = str(movement).zfill(2)
                movement_old_col = movement_new_col+"_2021"
                movement_diff_col = movement_new_col+"_diff"
                value_new = df.loc[i, movement_new_col]
                value_old = src_ws2.cell(row = row, column = col).value
                diff_ration = round((value_new - value_old)/value_old, 2) if value_old != 0 else 1.0
                heatmap_df.loc[i, movement_new_col] = diff_ration
                output_df.loc[i, movement_old_col] = value_old
                output_df.loc[i, movement_diff_col] = f"{int(diff_ration*100)}%"           
        i += 1
    
    src_wb.close()

    heatmap_df.set_index(["지점", "hm", "차종"], inplace=True)
    plt.figure(figsize=(10, 20))
    sns.heatmap(heatmap_df, annot=True, annot_kws={"size": 18})
    plt.savefig(output_fig_path)
    plt.close()
    print(f"Saved!!! >>> {output_fig_path}")

    sorted_col = list(output_df.columns[:3]) + sorted(output_df.columns[3:])
    output_df = output_df[sorted_col]
    output_df.to_excel(output_path, index=False, encoding="cp949")
    print(f"Saved!!! >>> {output_path}")

    diff_list = output_df[output_df["차종"] == "small"][movement_diff_col].apply(lambda x: float(x[:-1])/100)
    mean_diff = f"{diff_list.mean()*100:.2f}%"
    diff_list_include_large = output_df[movement_diff_col].apply(lambda x: float(x[:-1])/100)
    mean_diff_inclued_large = f"{diff_list_include_large.mean()*100:.2f}%"

    return mean_diff, mean_diff_inclued_large


def read_and_write_value(input_path: str, output_path: str, sheet1: dict, sheet2: dict) -> None:
    src_wb: Workbook
    src_wb = load_workbook_ignore_warning(input_path=input_path)

    src_ws1 = src_wb[sheet1["sheet_name"]]
    # mr = src_ws1.max_row
    # mc = src_ws1.max_column
    for row in sheet1["row_range1"]:
        for col in sheet1["col_range1"]:
            random_factor = get_random_value()
            old_value = src_ws1.cell(row = row, column = col).value
            new_value = int(round(old_value * random_factor))
            src_ws1.cell(row = row, column = col).value = new_value

    for row in sheet1["row_range2"]:
        for col in sheet1["col_range2"]:
            random_factor = get_random_value()
            old_value = src_ws1.cell(row = row, column = col).value
            new_value = int(round(old_value * random_factor))
            src_ws1.cell(row = row, column = col).value = new_value


    src_ws2 = src_wb[sheet2["sheet_name"]]
    for row in sheet2["row_range1"]:
        for col in sheet2["col_range1"]:
            random_factor = get_random_value()
            old_value = src_ws2.cell(row = row, column = col).value
            new_value = int(round(old_value * random_factor))
            src_ws2.cell(row = row, column = col).value = new_value


    src_wb.save(output_path)
    print(f"Saved!!! >>> {output_path}")
    src_wb.close()


# def col_split(row):
#     ymd, hms = row["time"].split(" ")
#     y, m, d = ymd.split("-")
#     h, m, s = hms.split(":")
#     hm = ":".join([h, m])
#     try: 
#         s, ms = s.split(".")
#     except:
#         ms = "00"
    
#     spot = row["spot"]
#     movement = row["movement"]
#     vtype = row["vehicle"]
#     traffic = 1
#     return ymd, d, h, m, s, ms, hm, spot, movement, vtype, traffic


def main() -> None:
    # txt_path = os.path.join(TXT_ROOT, TXT_FILE)
    # rename_dict = {
    # 0:"time",
    # 1:"spot",
    # 2:"movement",
    # 3:"speed",
    # 4:"vehicle"
    # }
    # df_raw = pd.read_csv(txt_path, header=None, encoding="utf-8")
    # df_raw = df_raw.rename(columns=rename_dict)
    # df_raw = df_raw.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    # col_list = ["연월일", "일", "h", "m", "s", "ms", "hm", "지점", "방향", "차종", "교통량"]
    # df = pd.DataFrame(columns=col_list)
    # df[col_list] = df_raw.apply(col_split, axis=1, result_type="expand")

    # fifteen_df = get_fifteen_df(TXT_FILE)

    ph_spot_move_dict = {str(key).zfill(2): val for key, val in PH_SPOT_MOVE_DICT.items()}

    now = datetime.now().strftime("%Y_%m_%d__%H_%M_%S")
    output_dir = os.path.join(OUTPUT_ROOT, now)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if EXCEL_FILE:
        input_name, ext = os.path.splitext(EXCEL_FILE)
        output_file = input_name + "_new" + ext
        input_path = os.path.join(INPUT_ROOT, EXCEL_FILE)
        output_path = os.path.join(output_dir, output_file)
        read_and_write_value(input_path, output_path, WRITE_SHEET1, WRITE_SHEET2)
    else:
        input_list = os.listdir(INPUT_ROOT)
        summary_df = pd.DataFrame()
        for input_file in input_list:
            input_name, ext = os.path.splitext(input_file)
            output_file = input_name + "_new" + ext
            # output_file = input_name + ext
            input_path = os.path.join(INPUT_ROOT, input_file)
            output_path = os.path.join(output_dir, output_file)
            read_and_write_value(input_path, output_path, WRITE_SHEET1, WRITE_SHEET2)
        #     spot = input_name[:2]
        #     if spot in list(ph_spot_move_dict.keys()):
        #         spotwise_df = fifteen_df.loc[fifteen_df["지점"] == spot]
        #         movement_list = ph_spot_move_dict[spot]
        #         output_file = input_name + "_compare" + ext
        #         output_path = os.path.join(output_dir, output_file)
        #         summary_df.loc[spot, ["평균증감율", "평균증감율(대형포함)"]] = \
        #             read_and_compare_value(spotwise_df, input_path, output_path, WRITE_SHEET1, WRITE_SHEET2, movement_list)
        # output_summary_file = "0_total_summary" + ext
        # output_summary_path = os.path.join(output_dir, output_summary_file)
        # summary_df.to_excel(output_summary_path, encoding="cp949")
        # print(f"Saved!!! >>> {output_summary_path}")
                


if __name__ == "__main__":
    main()


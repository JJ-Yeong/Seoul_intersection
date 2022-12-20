import os
import warnings
from openpyxl import Workbook as WB
from openpyxl import load_workbook


def load_workbook_ignore_warning(path: str, data_only: bool=False) -> WB:
    warnings.simplefilter("ignore")
    src_wb = load_workbook(filename=path, data_only=data_only)
    warnings.simplefilter("default")
    return src_wb

def read_and_write_value(ref_path: str, input_path: str, output_path: str) -> None:
    ref_wb: WB = load_workbook_ignore_warning(path=ref_path, data_only=True)
    input_wb: WB = load_workbook_ignore_warning(path=input_path)

    ref_ws = ref_wb["전년도교통량비교"]
    input_ws = input_wb["전년도교통량비교"]

    ref_2020 = ref_ws.cell(row=1, column=16).value   # 2020
    ref_2021 = ref_ws.cell(row=1, column=17).value   # 2021
    if ref_2020 == "2020년" and ref_2021 == "2021년":
        read_and_write_value.spot_list.append(os.path.basename(input_path).split(".x")[0])

        print()
        print("======path======\n")
        print("ref_path", ref_path)
        print("input_path", input_path)

        print("======ref year======")
        print("ref_2020", ref_2020)
        print("ref_2021", ref_2021)


        print("======ref 2021 value======")
        ref_2021_val1 = ref_ws.cell(row=2, column=17).value
        ref_2021_val2 = ref_ws.cell(row=3, column=17).value
        ref_2021_val3 = ref_ws.cell(row=4, column=17).value
        print("ref_2021_val1", ref_2021_val1)
        print("ref_2021_val2", ref_2021_val2)
        print("ref_2021_val3", ref_2021_val3)


        print("======input year======")
        input_2020 = input_ws.cell(row=1, column=16).value # 2020
        input_2021 = input_ws.cell(row=1, column=17).value # 2021
        input_ws.cell(row=1, column=16).value = "2021년"
        input_ws.cell(row=1, column=17).value = "2022년"
        print("input_2020", input_2020, "->", "2021년")
        print("input_2021", input_2021, "->", "2022년")


        print("======input 2020 value======")
        input_2020_val1 = input_ws.cell(row=2, column=16).value
        input_2020_val2 = input_ws.cell(row=3, column=16).value
        input_2020_val3 = input_ws.cell(row=4, column=16).value
        input_ws.cell(row=2, column=16).value = ref_2021_val1
        input_ws.cell(row=3, column=16).value = ref_2021_val2
        input_ws.cell(row=4, column=16).value = ref_2021_val3
        print("input_2020_val1", input_2020_val1, "->", ref_2021_val1)
        print("input_2020_val2", input_2020_val2, "->", ref_2021_val2)
        print("input_2020_val3", input_2020_val3, "->", ref_2021_val3)


        print("======input 2022 value======")
        input_2022_val1 = input_ws.cell(row=2, column=17).value
        input_2022_val2 = input_ws.cell(row=3, column=17).value
        input_2022_val3 = input_ws.cell(row=4, column=17).value
        print("input_2022_val1", input_2022_val1)
        print("input_2022_val2", input_2022_val2)
        print("input_2022_val3", input_2022_val3)

        input_wb.save(output_path)
        print(f"Saved!!! >>> {output_path}")
        input_wb.close()


    # mr = src_ws1.max_row
    # mc = src_ws1.max_column
    # for row in sheet1["row_range1"]:
    #     for col in sheet1["col_range1"]:
    #         src_ws1.cell(row = row, column = col).value

    # for row in sheet1["row_range2"]:
    #     for col in sheet1["col_range2"]:
    #         src_ws1.cell(row = row, column = col).value


    # src_ws2 = src_wb[sheet2["sheet_name"]]
    # for row in sheet2["row_range1"]:
    #     for col in sheet2["col_range1"]:
    #         src_ws2.cell(row = row, column = col).value



if __name__ == "__main__":
    
    REF_ROOT = "excel_input/total" 
    INPUT_ROOT = "221109_수정전"
    OUPUT_ROOT = "221109_수정후"
    ref_list = sorted(os.listdir(REF_ROOT))
    input_list = sorted(os.listdir(INPUT_ROOT))
    for ref_name, input_name in zip(ref_list, input_list):
        ref_path = os.path.join(REF_ROOT, ref_name)
        input_path = os.path.join(INPUT_ROOT, input_name)
        output_path = os.path.join(OUPUT_ROOT, input_name)
        read_and_write_value.spot_list = []
        read_and_write_value(ref_path, input_path, output_path)
    
    print( *read_and_write_value.spot_list, sep="\n")